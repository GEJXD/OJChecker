import requests
from openpyxl import load_workbook
from urllib.parse import quote

import time
import logging
import json
from pathlib import Path

# vjudge headers
headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.108 Safari/537.36'
}

configFileName = {'vj': 'vjudge', 'nc': 'nowcoder', 'jsk': 'jisuanke'}


def get_user(OJ: str):
    logging.info(f"Loading {OJ} user infomation")

    path = Path('./config/')
    path = path.joinpath(configFileName[OJ] + '.xlsx')

    if not path.is_file():
        logging.exception(f"The file {path} dose not exist")

    workbook = load_workbook(path)
    sheet = workbook[workbook.sheetnames[0]]

    rows = sheet.rows
    data = {}

    for id, row in enumerate(rows):
        if row[0].value is None:
            break
        if id == 0:
            continue

        who = row[0].value
        user = row[1].value
        data[who] = str(user)

    return data


def login(username, password):
    logUrl = "https://vjudge.net/user/login"
    data = {
            'username': username,
            'password': password
    }

    seesion = requests.session()
    while True:
        try:
            response = seesion.post(logUrl, headers=headers, data=data)
            break
        except Exception as err:
            logging.error(f'Error {err}')
            logging.error(f'Retrying')
    
    logging.info(f"Getting Login Cookies Succeeded")
    return response.cookies


def get_contest(OJ: str, contest_id: int, cookies = None):
    url = ''
    if OJ == 'vj':
        url = f"https://vjudge.net/contest/rank/single/{contest_id}"
    elif OJ == 'nc':
        url = f"https://ac.nowcoder.com/acm/contest/rank/submit-list?token=&currentContestId={contest_id}&contestList={contest_id}&_={int(time.time())}"
    while True:
        try:
            if cookies is None:
                res = requests.get(url)
            else:
                res = requests.get(url, cookies=cookies)
            break
        except Exception as err:
            logging.error(f'Error {err}')
            logging.error(f'Retrying')
    logging.info(f"Getting Submit Messages Succeeded")
    return res.text

def fliter(data, users):
    user_names = []
    for key, val in users.items():
        user_names.append(val)

    contest = json.loads(data)
    problemData = contest['data']['problemData']

    problems = {}
    for problem in problemData:
        problems[problem['problemId']] = problem['index']

    submitData = contest['data']['submitDataList'][0]
    basicInfo = submitData['basicInfo']
    length = (basicInfo['endTime'] - basicInfo['startTime']) // 1000
    signUpUsers = submitData['signUpUsers']

    user_uid = {}
    for i in signUpUsers:
        if i['name'] in user_names:
            user_uid[i['uid']] = i['name']
    
    contest_data = {}
    submissions = submitData['submissions']

    for submission in submissions:
        if submission['uid'] not in user_uid:
            continue
        who = user_uid[submission['uid']]
        if who not in contest_data:
            contest_data[who] = {}
            contest_data[who]['solved'] = []
            contest_data[who]['aSolved'] = []
            contest_data[who]['totTime'] = 0
            contest_data[who]['attempted'] = False

        if submission['submitTime'] <= basicInfo['endTime']:
            contest_data[who]['attempted'] = True

        which = ord(problems[submission['problemId']]) - ord('A')

        if submission['status'] == 5 and which not in contest_data[who]['solved']:
            if submission['submitTime'] <= basicInfo['endTime']:
                contest_data[who]['solved'].append(which)
                if which not in contest_data[who]:
                    contest_data[who][which] = 0
                contest_data[who]['totTime'] += (submission['submitTime'] - basicInfo['startTime']) // 1000 + contest_data[who][which] * 20 * 60
        elif submission['status'] != 5 and which not in contest_data[who]['solved']:
            if which not in contest_data[who]:
                contest_data[who][which] = 0
            contest_data[who][which] += 1
    contest_data_f = {}
    for key, val in sorted(contest_data.items(), key=lambda item: len(item[1]['solved'])):
        contest_data_f[key] = val
    return contest_data_f, basicInfo['endTime']
